import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except ImportError:
    install('openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "마진계산기"

# Define styles
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

bold_font = Font(bold=True)
red_bold_font = Font(bold=True, color="FF0000")
green_font = Font(color="375623")
green_bold_font = Font(color="375623", bold=True)
blue_bold_font = Font(color="000000", bold=True)

center_alignment = Alignment(horizontal="center", vertical="center")
right_alignment = Alignment(horizontal="right", vertical="center")

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Set column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 15

# Data and structure
data = [
    ["구분", "금액", "산출근거", "비고"],
    ["소비자 판매가", 13920, "온라인 최저가 고정", "620g*3묶음"],
    ["쿠팡 카테고리 수수료(11.99%)", "=ROUND(B2*0.1199, 0)", "소비자 판매가 X 11.99%", ""],
    ["로켓그로스 물류 고정비", 2420, "입고 + 출고 + 배송(소형 기준 합산)", ""],
    ["쿠팡 판매 광고비(11%)", "=ROUND(B2*0.11, 0)", "최저가 노출 및 매출 유지를 위한 최소 광고비", ""],
    ["물류 센터 보관료 및 리스트 비용(4%)", "=ROUND(B2*0.04, 0)", "반품 처리, 유통기한/재고 파손, 라벨비 작업비", ""],
    ["쿠팡 공제 후 최종 정산금액", "=B2-SUM(B3:B6)", "2-(3+4+5+6)", ""],
    ["아이비 마진(25%)", "=ROUND(B2*0.25, 0)", "사업 유지를 위한 최소 마진(목표치)", ""],
    ["3개 세트 총 목표 원가", "=B7-B8", "7-8", ""],
    ["개당 최대 목표 공급가", "=ROUND(B9/3, 0)", "9/3", ""]
]

for row_idx, row_data in enumerate(data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        
        # Default alignment
        if col_idx == 2:
            cell.alignment = right_alignment
            if row_idx > 1:
                cell.number_format = '#,##0'
        else:
            cell.alignment = center_alignment

        # Styling
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = bold_font
        elif row_idx == 2:
            if col_idx <= 3:
                cell.fill = green_fill
            if col_idx == 1 or col_idx == 3:
                cell.font = green_font
            if col_idx == 2:
                cell.font = green_bold_font
        elif row_idx == 7:
            if col_idx <= 3:
                cell.fill = blue_fill
            if col_idx <= 3:
                cell.font = blue_bold_font
        elif row_idx == 8:
            if col_idx <= 3:
                cell.fill = yellow_fill
            if col_idx <= 3:
                cell.font = red_bold_font
        elif row_idx == 10:
            if col_idx <= 3:
                cell.fill = yellow_fill
            if col_idx <= 3:
                cell.font = red_bold_font

wb.save("d:/NATAS Harnes-menu/margin_calculator.xlsx")
print("Excel file created at d:/NATAS Harnes-menu/margin_calculator.xlsx")
